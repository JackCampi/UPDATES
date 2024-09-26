from sqlalchemy.orm import Session
from ..model.insert_conf import InsertConf
from ..data_management.data_connector import read_not_found, build_not_found_report_path
import pandas as pd
from .db_controller import get_pro_name
from .utils.pac_utils import get_pac_text

from openpyxl.styles import PatternFill, NamedStyle, Alignment, Font, Border, Side
from openpyxl.utils.cell import get_column_letter


def not_found_report(name: str, insert: InsertConf, db: Session):
    csv = read_not_found(insert, name)

    noms = []
    for index in csv.index:
        pro_nom = get_pro_name(db, csv['COD_PROGRAMA'][index])
        noms.append(pro_nom)

    csv['PROGRAMA'] = noms

    title, pac = get_pac_text(insert)
    pacs = [pac] * len(csv.index)

    csv.insert(0, title, pacs)

    path = build_not_found_report_path(insert, name)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        csv.to_excel(writer, sheet_name="REPORT", index=False)
        workbook = writer.book
        worksheet = workbook["REPORT"]
        for column in worksheet.columns:
            l = [len(str(cell.value)) for cell in column]
            max_length = max(l)
            adjusted_width = (max_length + 4) * 1.0
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        for row in worksheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(fill_type="solid", start_color="E49EDD", end_color="E49EDD")
    return{
        "FINITI" : True
    }
